import json
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, update_session_auth_hash, logout
from .models import (
    TipoSocio, Socio, CuentaBancaria, MetodoPago, Prestamo, Pago, 
    Bingo, Ahorro, Jugador, PartidaBingo, Carton, CartonPartidaBingo, 
    PlataformaJuego, SesionJuego, Regalo, AporteSemanal, ConfiguracionWeb, UnidadMonetaria, MensajeChat
)
from .services import generar_matriz_bingo, generar_lote_cartones, actualizar_socio_y_credenciales, actualizar_jugador_y_credenciales, actualizar_avatar_perfil, validar_carton_hibrido
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Sum, Q, ProtectedError
from django.contrib.auth.decorators import login_required
from datetime import datetime, date, timedelta
from django.db import transaction
from django.http import JsonResponse, HttpResponse
import uuid
from django.utils import timezone
import random
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
from .tasks import fabricar_cartones_maestros_task
from django.template.loader import get_template
from xhtml2pdf import pisa
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from django.shortcuts import render
from decimal import Decimal
from django.shortcuts import render, redirect




def inicio(request):
    preguntar_jugador = request.session.pop('preguntar_jugador', False)
    es_jugador = False
    mostrar_promo_socio = False

    if request.user.is_authenticated and not request.user.is_staff:
        jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
        if jugador:
            es_jugador = True
            # LÓGICA DE ÚNICA VEZ: Si es jugador, NO es socio, y no ha visto la promo en esta sesión
            if not jugador.idsocio and not request.session.get('promo_socio_visto', False):
                mostrar_promo_socio = True
                request.session['promo_socio_visto'] = True

    config_web = ConfiguracionWeb.objects.first()
    
    # 1. Recuperamos los bingos activos
    bingos_activos = Bingo.objects.filter(
        estadobingo__in=['Programado', 'En Curso']
    ).select_related('idunidadmonetaria').order_by('fechaprogramadabingo')  
    
    # ================================================================
    # CÓDIGO LIMPIO: Django ya sabe que estamos en America/Guayaquil
    # ================================================================
    ahora = timezone.now() 
    
    for b in bingos_activos:
        if b.fechaprogramadabingo:
            # Restamos 30 minutos limpiamente
            hora_apertura = b.fechaprogramadabingo - timedelta(minutes=30)
            
            # Buscamos si hay una ronda lista
            partida_activa = PartidaBingo.objects.filter(
                idbingo=b,
                estadopartida__in=['Programada', 'En Juego']
            ).order_by('idpartidabingo').first()
            
            # Comparación nativa y elegante
            if ahora >= hora_apertura and partida_activa:
                b.sala_abierta = True
                b.id_partida_a_entrar = partida_activa.idpartidabingo
            else:
                b.sala_abierta = False
        else:
            b.sala_abierta = False

    contexto = {
        'preguntar_jugador': preguntar_jugador,
        'es_jugador': es_jugador,
        'config_web': config_web,
        'bingos_activos': bingos_activos,
        'mostrar_promo_socio': mostrar_promo_socio,
    }
    return render(request, 'comunes/inicio.html', contexto)

@login_required
def dashboard(request):
    if not request.user.is_staff:
        messages.error(request, "Acceso exclusivo para el personal de administración.")
        return redirect('inicio')

    if request.method == 'POST':
        action = request.POST.get('action')
        try:
            if action == 'crear_tiposocio':
                TipoSocio.objects.create(nombretiposocio=request.POST.get('nombretiposocio'), roltiposocio=request.POST.get('roltiposocio'), descripciondetiposocio=request.POST.get('descripciondetiposocio'))
                messages.success(request, "Tipo de Socio creado correctamente.")
            elif action == 'eliminar_tiposocio':
                TipoSocio.objects.get(idtiposocio=request.POST.get('id_tipo')).delete()
                messages.success(request, "Tipo de Socio eliminado.")
            elif action == 'crear_plataforma':
                estado_plat = True if request.POST.get('estadoplataforma') == 'on' else False
                PlataformaJuego.objects.create(nombreplataforma=request.POST.get('nombreplataforma'), urlplataforma=request.POST.get('urlplataforma'), descripcionplataforma=request.POST.get('descripcionplataforma'), contactoplataforma=request.POST.get('contactoplataforma'), estadoplataforma=estado_plat, fechaadquisicionlicencia=request.POST.get('fechaadquisicionlicencia') or None, fechavencimientolicencia=request.POST.get('fechavencimientolicencia') or None, logoplataforma=request.FILES.get('logoplataforma'))
                messages.success(request, "Plataforma de Juego registrada con éxito.")
            elif action == 'eliminar_plataforma':
                PlataformaJuego.objects.get(idplataformajuego=request.POST.get('id_plataforma')).delete()
                messages.success(request, "Plataforma eliminada del sistema.")
            elif action == 'crear_bingo':
                unidad = get_object_or_404(UnidadMonetaria, idunidadmonetaria=request.POST.get('idunidadmonetaria'))
                Bingo.objects.create(idunidadmonetaria=unidad, titulobingo=request.POST.get('titulobingo'), fechaprogramadabingo=request.POST.get('fechaprogramadabingo'), tipobingo=request.POST.get('tipobingo'), lugarbingo=request.POST.get('lugarbingo'), urlsesionbingo=request.POST.get('urlsesionbingo'), preciocarton=request.POST.get('preciocarton'), premiomayor=request.POST.get('premiomayor'), descripcionpremiomayor=request.POST.get('descripcionpremiomayor'), estadobingo=request.POST.get('estadobingo'), descripcionpremios=request.POST.get('descripcionpremios'), rutaimagenpremiomayor=request.FILES.get('rutaimagenpremiomayor'), urlvideopromocional=request.FILES.get('urlvideopromocional'))
                messages.success(request, "¡Jornada de Bingo creada exitosamente!")
                
            elif action == 'editar_bingo':
                bingo = Bingo.objects.get(idbingo=request.POST.get('id_bingo'))
                bingo.idunidadmonetaria = get_object_or_404(UnidadMonetaria, idunidadmonetaria=request.POST.get('idunidadmonetaria'))
                bingo.titulobingo = request.POST.get('titulobingo')
                bingo.preciocarton = request.POST.get('preciocarton')
                bingo.premiomayor = request.POST.get('premiomayor')
                bingo.descripcionpremiomayor = request.POST.get('descripcionpremiomayor')
                bingo.descripcionpremios = request.POST.get('descripcionpremios')
                if request.POST.get('fechaprogramadabingo'): bingo.fechaprogramadabingo = request.POST.get('fechaprogramadabingo')
                bingo.tipobingo = request.POST.get('tipobingo')
                bingo.lugarbingo = request.POST.get('lugarbingo')
                bingo.urlsesionbingo = request.POST.get('urlsesionbingo')
                estado_anterior = bingo.estadobingo
                nuevo_estado = request.POST.get('estadobingo')
                bingo.estadobingo = nuevo_estado
                if 'rutaimagenpremiomayor' in request.FILES: bingo.rutaimagenpremiomayor = request.FILES['rutaimagenpremiomayor']
                if 'urlvideopromocional' in request.FILES: bingo.urlvideopromocional = request.FILES['urlvideopromocional']
                bingo.save()
                
                # ==========================================================
                # ÁRBITRO DIGITAL: DISPARO INICIAL (Variables Corregidas)
                # ==========================================================
                if nuevo_estado == 'En Curso' and estado_anterior != 'En Curso':
                    primera_partida = PartidaBingo.objects.filter(idbingo=bingo).order_by('idpartidabingo').first()
                    if primera_partida and primera_partida.estadopartida == 'Programada':
                        primera_partida.estadopartida = 'En Juego'
                        primera_partida.horainiciopartida = timezone.now() # CORREGIDO A horainiciopartida
                        primera_partida.save()
                        messages.success(request, "¡Bingo iniciado! La primera ronda ha comenzado automáticamente.")
                # ==========================================================

                if nuevo_estado == 'Finalizado' and estado_anterior != 'Finalizado':
                    cartones_temporales = CartonPartidaBingo.objects.filter(idpartida__idbingo=bingo, idcarton__esmaestro=False).values_list('idcarton', flat=True)
                    ids_a_borrar = list(set(cartones_temporales))
                    if ids_a_borrar:
                        CartonPartidaBingo.objects.filter(idpartida__idbingo=bingo, idcarton__esmaestro=False).delete()
                        Carton.objects.filter(idcarton__in=ids_a_borrar).delete()
                        messages.success(request, f"¡Bingo Finalizado! El sistema ha autodestruido {len(ids_a_borrar)} cartones temporales.")
                    else:
                        messages.success(request, "Jornada de Bingo actualizada y Finalizada correctamente.")
                else:
                    messages.success(request, "Jornada de Bingo actualizada correctamente.")
                
            elif action == 'eliminar_bingo':
                Bingo.objects.get(idbingo=request.POST.get('id_bingo')).delete()
                messages.success(request, "Jornada de Bingo eliminada por completo.")
                
            elif action == 'crear_partida':
                bingo_obj = Bingo.objects.get(idbingo=request.POST.get('idbingo'))
                
                # FIX FASE 2: LÓGICA DEL PREMIO MAYOR ÚNICO
                es_pozo_mayor = request.POST.get('es_pozo_mayor') == 'on'
                
                if es_pozo_mayor:
                    valor_premio = 0
                    premio_material = '[POZO_MAYOR]' # Etiqueta secreta para el motor de pagos
                else:
                    valor_premio = request.POST.get('valorpremio')
                    premio_material = request.POST.get('premiomaterial')
                    if not valor_premio or str(valor_premio).strip() == '': valor_premio = 0
                    if not premio_material or str(premio_material).strip() == '': premio_material = 'Ninguno'
                
                PartidaBingo.objects.create(
                    idbingo=bingo_obj, 
                    nombreronda=request.POST.get('nombreronda'), 
                    modalidad_victoria=request.POST.get('modalidad_victoria', 'Tabla Llena'),
                    valorpremio=valor_premio, 
                    premiomaterial=premio_material, 
                    estadopartida='Programada', 
                    bolascantadas='', 
                    ultimabola=0 
                )
                
                if es_pozo_mayor:
                    messages.success(request, f"¡Ronda '{request.POST.get('nombreronda')}' aperturada! Jugarán por el POZO MAYOR de ${bingo_obj.premiomayor}.")
                else:
                    messages.success(request, f"¡Ronda '{request.POST.get('nombreronda')}' aperturada con modalidad {request.POST.get('modalidad_victoria')}!")
                
            elif action == 'eliminar_partida':
                PartidaBingo.objects.get(idpartidabingo=request.POST.get('id_partida')).delete()
                messages.success(request, "Ronda eliminada de forma segura.")
            # =======================================================
            # NUEVO: LOGÍSTICA DE ENTREGA DE PREMIOS FÍSICOS
            # =======================================================
            elif action == 'entregar_premio_fisico':
                partida = PartidaBingo.objects.get(idpartidabingo=request.POST.get('id_partida'))
                partida.estadopremiomaterial = 'Entregado'
                partida.save()
                messages.success(request, f"¡Excelente! El premio físico de la ronda '{partida.nombreronda}' ha sido marcado como ENTREGADO.")
            # =======================================================
            elif action == 'editar_configuracion':
                config, created = ConfiguracionWeb.objects.get_or_create(idconfiguracion=1)
                config.titulosobrenosotros = request.POST.get('titulosobrenosotros', config.titulosobrenosotros)
                config.descripcionsobrenosotros = request.POST.get('descripcionsobrenosotros', config.descripcionsobrenosotros)
                config.numerowhatsapp = request.POST.get('numerowhatsapp', config.numerowhatsapp)
                config.enlaceinstagram = request.POST.get('enlaceinstagram', config.enlaceinstagram)
                config.enlacefacebook = request.POST.get('enlacefacebook', config.enlacefacebook)
                if 'imagenpromocional' in request.FILES: config.imagenpromocional = request.FILES['imagenpromocional']
                config.save()
                messages.success(request, "Configuración del sitio web actualizada correctamente.")
            elif action == 'generar_cartones':
                cantidad = int(request.POST.get('cantidad_cartones', 0))
                if cantidad > 0:
                    lote = generar_lote_cartones(cantidad)
                    cartones_db = [Carton(codigocarton=c['codigo'], matriznumeros=c['matriz'], esmaestro=True) for c in lote]
                    Carton.objects.bulk_create(cartones_db)
                    fabricar_cartones_maestros_task.delay(cantidad)
                    messages.success(request, f"¡Orden enviada a la fábrica! Se están estampando {cantidad} cartones RNG en segundo plano.")
            elif action == 'eliminar_carton':
                Carton.objects.get(idcarton=request.POST.get('id_carton')).delete()
                messages.success(request, "Cartón retirado del inventario general.")
            elif action == 'editar_socio':
                actualizar_socio_y_credenciales(request.POST.get('id_socio'), request.POST.get('cedula'), request.POST.get('nombres'), request.POST.get('apellidos'), request.POST.get('telefono'), request.POST.get('estado'), request.POST.get('id_tipo_socio'), request.POST.get('password_nueva'))
                messages.success(request, f"Perfil del socio actualizado correctamente.")
            elif action == 'editar_jugador':
                actualizar_jugador_y_credenciales(request.POST.get('id_jugador'), request.POST.get('alias'), request.POST.get('cedula'), request.POST.get('correo'), request.POST.get('estado'), request.POST.get('password_nueva'))
                messages.success(request, f"Perfil del jugador actualizado correctamente.")
            elif action == 'crear_moneda':
                estado = True if request.POST.get('estadomoneda') == 'on' else False
                UnidadMonetaria.objects.create(
                    nombremoneda=request.POST.get('nombremoneda'),
                    tipomoneda=request.POST.get('tipomoneda'),
                    simbolomoneda=request.POST.get('simbolomoneda'),
                    tasaconversionmoneda=request.POST.get('tasaconversionmoneda'),
                    estadomoneda=estado
                )
                messages.success(request, "Nueva unidad monetaria registrada con éxito.")
                
            elif action == 'editar_moneda':
                moneda = UnidadMonetaria.objects.get(idunidadmonetaria=request.POST.get('id_moneda'))
                moneda.nombremoneda = request.POST.get('nombremoneda')
                moneda.tipomoneda = request.POST.get('tipomoneda')
                moneda.simbolomoneda = request.POST.get('simbolomoneda')
                moneda.tasaconversionmoneda = request.POST.get('tasaconversionmoneda')
                moneda.estadomoneda = True if request.POST.get('estadomoneda') == 'on' else False
                moneda.save()
                messages.success(request, "Divisa actualizada correctamente.")

            elif action == 'crear_ahorro':
                socio = get_object_or_404(Socio, idsocio=request.POST.get('id_socio'))
                monto_ahorro = request.POST.get('montoahorro')
                
                Ahorro.objects.create(
                    idsocio=socio,
                    montoahorro=monto_ahorro,
                    fechaahorro=timezone.now(),
                    estadoahorro='Acreditado'
                )
                messages.success(request, f"Se ha registrado un ahorro de ${monto_ahorro} para el socio {socio.primernombresocio}.")

            elif action == 'crear_prestamo':
                socio = get_object_or_404(Socio, idsocio=request.POST.get('id_socio'))
                monto_solicitado = request.POST.get('montoprestamosolicitado')
                interes = request.POST.get('tasainteres', 0)
                
                # Cálculo simple del total a pagar (Monto + (Monto * interes / 100))
                monto_total = float(monto_solicitado) + (float(monto_solicitado) * float(interes) / 100)
                
                Prestamo.objects.create(
                    idsocio=socio,
                    montoprestamosolicitado=monto_solicitado,
                    montototalpagar=monto_total,
                    saldopendiente=monto_total,
                    fechasolicitud=timezone.now(),
                    estadoprestamo='Aprobado'
                )
                messages.success(request, f"Préstamo aprobado para {socio.primernombresocio} por un total a pagar de ${monto_total}.")

            elif action == 'registrar_pago_prestamo':
                prestamo = get_object_or_404(Prestamo, idprestamo=request.POST.get('id_prestamo'))
                monto_abonado = request.POST.get('monto_abono')
                metodo_pago = get_object_or_404(MetodoPago, idmetodopago=request.POST.get('id_metodo_pago'))
                
                # Restar el saldo
                prestamo.saldopendiente = float(prestamo.saldopendiente) - float(monto_abonado)
                
                # Si pagó todo, se liquida
                if prestamo.saldopendiente <= 0:
                    prestamo.saldopendiente = 0
                    prestamo.estadoprestamo = 'Liquidado'
                prestamo.save()
                
                # Guardar el registro del pago
                Pago.objects.create(
                    idprestamo=prestamo,
                    idmetodopago=metodo_pago,
                    montopago=monto_abonado,
                    fechapago=timezone.now(),
                    estadopago='Completado'
                )
                messages.success(request, f"Abono de ${monto_abonado} registrado exitosamente al préstamo.")
                # =======================================================
            # APROBAR O RECHAZAR CRÉDITOS DESDE EL DASHBOARD
            # =======================================================
            elif action == 'aprobar_prestamo':
                id_prestamo = request.POST.get('id_prestamo')
                prestamo = get_object_or_404(Prestamo, idprestamo=id_prestamo)
                
                # Cambiamos el estado a Aprobado para activar el crédito
                prestamo.estadoprestamo = 'Aprobado'
                prestamo.save()
                messages.success(request, f"El crédito #{prestamo.idprestamo} para {prestamo.idsocio.primernombresocio} ha sido Aprobado.")

            elif action == 'rechazar_prestamo':
                id_prestamo = request.POST.get('id_prestamo')
                prestamo = get_object_or_404(Prestamo, idprestamo=id_prestamo)
                
                # Como tu modelo no tiene el estado 'Rechazado', eliminamos la solicitud denegada
                prestamo.delete()
                messages.warning(request, "La solicitud de crédito ha sido rechazada y eliminada del sistema.")
            # =======================================================
            # NUEVO: VALIDAR O RECHAZAR PAGOS (AMORTIZACIONES)
            # =======================================================
            elif action == 'validar_pago':
                id_pago = request.POST.get('id_pago')
                pago_obj = get_object_or_404(Pago, idpago=id_pago)
                
                if pago_obj.estadopago == 'Pendiente':
                    # Descontar el dinero del saldo pendiente del préstamo asociado
                    prestamo = pago_obj.idprestamo
                    prestamo.saldopendiente = prestamo.saldopendiente - pago_obj.montopagado
                    
                    # Si el saldo llega a cero o menos, el préstamo queda liquidado
                    if prestamo.saldopendiente <= 0:
                        prestamo.saldopendiente = 0
                        prestamo.estadoprestamo = 'Liquidado'
                    prestamo.save()

                    # Actualizar el estado del pago a Validado
                    pago_obj.estadopago = 'Validado'
                    pago_obj.fechaconfirmacionadmin = timezone.now()
                    pago_obj.save()
                    
                    messages.success(request, f"¡Pago de ${pago_obj.montopagado} validado! El saldo del crédito ha sido descontado correctamente.")
            
            elif action == 'rechazar_pago':
                id_pago = request.POST.get('id_pago')
                pago_obj = get_object_or_404(Pago, idpago=id_pago)
                
                if pago_obj.estadopago == 'Pendiente':
                    pago_obj.estadopago = 'Rechazado'
                    pago_obj.fechaconfirmacionadmin = timezone.now()
                    pago_obj.save()
                    messages.warning(request, "El abono reportado ha sido rechazado.")
            elif action == 'crear_metodo_pago':
                nombre = request.POST.get('nombremetodopago')
                descripcion = request.POST.get('descripcionmetodopago')
                url_cuenta = request.POST.get('urlmetodopago')
                estado = request.POST.get('estadometodopago')
                
                try:
                    MetodoPago.objects.create(
                        nombremetodopago=nombre,
                        descripcionmetodopago=descripcion,
                        urlmetodopago=url_cuenta,
                        estadometodopago=estado
                    )
                    messages.success(request, "Cuenta de destino registrada exitosamente.")
                except Exception as e:
                    messages.error(request, f"Error al registrar la cuenta: {str(e)}")
            elif action == 'validar_ahorro':
                id_ahorro = request.POST.get('id_ahorro')
                if id_ahorro:
                    ahorro_obj = get_object_or_404(Ahorro, idahorro=id_ahorro)
                    ahorro_obj.estadoahorro = 'Acreditado'
                    ahorro_obj.save()
                    messages.success(request, f"¡Depósito por ${ahorro_obj.montoahorro} verificado y acreditado a la cuenta del socio!")
                return redirect('dashboard')

            elif action == 'rechazar_ahorro':
                id_ahorro = request.POST.get('id_ahorro')
                if id_ahorro:
                    ahorro_obj = get_object_or_404(Ahorro, idahorro=id_ahorro)
                    ahorro_obj.estadoahorro = 'Rechazado'
                    ahorro_obj.save()
                    messages.warning(request, "El reporte de ahorro fue rechazado por inconsistencias en el comprobante.")
                return redirect('dashboard')
            elif action == 'eliminar_moneda':
                UnidadMonetaria.objects.get(idunidadmonetaria=request.POST.get('id_moneda')).delete()
                messages.success(request, "Divisa eliminada del sistema.")
        except ProtectedError:
            messages.error(request, "⚠️ ERROR: No puedes eliminar este registro porque hay usuarios o datos vinculados a él.")
        except Exception as e:
            messages.error(request, f"Error en la operación: {str(e)}")
        return redirect('dashboard')
            

    contexto = {
        'total_socios': Socio.objects.count(), 'total_jugadores': Jugador.objects.count(), 'deuda_calle': Prestamo.objects.exclude(estadoprestamo='Liquidado').aggregate(total=Sum('saldopendiente'))['total'] or 0.00,
        'bingos_activos': Bingo.objects.exclude(estadobingo__in=['Finalizado', 'Cancelado']).count(), 'tipos_socio': TipoSocio.objects.all(),
        'socios': Socio.objects.all().order_by('-idsocio')[:50], 'accounts': CuentaBancaria.objects.all().select_related('idsocio'),
        'jugadores': Jugador.objects.all().order_by('-idjugador')[:50], 'prestamos': Prestamo.objects.all().order_by('-fechasolicitud')[:30],
        'pagos': Pago.objects.all().order_by('-fechapago')[:30], 'metodos_pago': MetodoPago.objects.all(),
        'ahorros': Ahorro.objects.all().order_by('-fechaahorro')[:30], 'aportes_semanales': AporteSemanal.objects.all().order_by('-fechaplanificadadada')[:30],
        'bingos': Bingo.objects.all().order_by('-fechaprogramadabingo'), 'partidas': PartidaBingo.objects.all(),
        'regalos': Regalo.objects.all(), 'cartones': Carton.objects.all().order_by('-idcarton')[:50],
        'cartones_en_juego': CartonPartidaBingo.objects.all()[:50], 'plataformas': PlataformaJuego.objects.all(),
        'sesiones_monitoreo': SesionJuego.objects.all().order_by('-fechainiciosesion')[:30], 'config_web': ConfiguracionWeb.objects.first(),
        'unidades_monetarias': UnidadMonetaria.objects.filter(estadomoneda=True),
        'todas_monedas': UnidadMonetaria.objects.all(),
        'metodos_pago': MetodoPago.objects.all().order_by('-idmetodopago'),
    }
    # Agrega esto al final de tus variables de contexto
    contexto['bingos_con_pozo'] = list(PartidaBingo.objects.filter(premiomaterial='[POZO_MAYOR]').values_list('idbingo_id', flat=True))
    
    return render(request, 'administrador/dashboard.html', contexto)

# Rutas Comunes y Cuentas
def bingo_publico(request):
    # Traemos los bingos que están por jugarse o en vivo (Para vender/promocionar)
    bingos_activos = Bingo.objects.filter(
        estadobingo__in=['Programado', 'En Curso']
    ).order_by('fechaprogramadabingo')

    # Traemos los bingos que ya terminaron (Historial)
    bingos_pasados = Bingo.objects.filter(
        estadobingo__in=['Finalizado', 'Cancelado']
    ).order_by('-fechaprogramadabingo') # Ordenados del más reciente al más antiguo

    # ================================================================
    # LÓGICA DE SALA DE ESPERA: Sincronizada con el inicio
    # ================================================================
    ahora = timezone.now() 
    
    for b in bingos_activos:
        if b.fechaprogramadabingo:
            hora_apertura = b.fechaprogramadabingo - timedelta(minutes=30)
            partida_activa = PartidaBingo.objects.filter(
                idbingo=b,
                estadopartida__in=['Programada', 'En Juego']
            ).order_by('idpartidabingo').first()
            
            if ahora >= hora_apertura and partida_activa:
                b.sala_abierta = True
                b.id_partida_a_entrar = partida_activa.idpartidabingo
            else:
                b.sala_abierta = False
        else:
            b.sala_abierta = False

    contexto = {
        'bingos_activos': bingos_activos,
        'bingos_pasados': bingos_pasados,
        'unidad_monetaria': UnidadMonetaria.objects.first()
    }
    return render(request, 'comunes/bingo.html', contexto)

@login_required
def cuenta_bancaria(request):
    """
    Vista para que el Socio registre y elimine sus cuentas bancarias para recibir depósitos.
    """
    socio = Socio.objects.filter(cisocio=request.user.username).first()
    if not socio:
        messages.warning(request, "Debes ser Socio para gestionar cuentas bancarias.")
        return redirect('inicio')

    # Obtener las cuentas vinculadas al socio
    cuentas = CuentaBancaria.objects.filter(idsocio=socio)

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'agregar_cuenta':
            CuentaBancaria.objects.create(
                idsocio=socio,
                bancocuentabancaria=request.POST.get('banco'),
                tipocuentabancaria=request.POST.get('tipo_cuenta'),
                numerocuentabancaria=request.POST.get('numero_cuenta'),
                titularcuentabancaria=request.POST.get('titular'),
                estadocuentabancaria='Activa'
            )
            messages.success(request, "Cuenta bancaria agregada exitosamente a tu perfil.")
            
        elif action == 'eliminar_cuenta':
            id_cuenta = request.POST.get('id_cuenta')
            # Nos aseguramos de que solo pueda eliminar SU propia cuenta
            cuenta = CuentaBancaria.objects.filter(idcuentabancaria=id_cuenta, idsocio=socio).first()
            if cuenta:
                cuenta.delete()
                messages.success(request, "La cuenta bancaria ha sido eliminada.")
        
        return redirect('cuenta_bancaria')

    contexto = {
        'socio': socio,
        'cuentas': cuentas
    }
    return render(request, 'cuentas/cuenta_bancaria.html', contexto)

@login_required
def ahorro(request):
    """
    Vista para que el Socio vea su libreta de ahorros, reporte depósitos y solicite retiros.
    """
    socio = Socio.objects.filter(cisocio=request.user.username).first()
    if not socio:
        messages.warning(request, "Debes ser Socio de la cooperativa para acceder a la libreta de ahorros.")
        return redirect('inicio')

    historial_ahorros = Ahorro.objects.filter(idsocio=socio).order_by('-fechaahorro')
    total_ahorrado = historial_ahorros.filter(estadoahorro='Acreditado').aggregate(total=Sum('montoahorro'))['total'] or Decimal('0.00')
    
    # NUEVO: Obtenemos las cuentas recaudadoras activas
    metodos_activos = MetodoPago.objects.filter(estadometodopago='Activo')

    if request.method == 'POST':
        action = request.POST.get('action')
        
        # 1. LOGICA PARA REPORTAR UN NUEVO AHORRO
        if action == 'registrar_ahorro':
            monto_ahorro = request.POST.get('monto_ahorro')
            id_metodo = request.POST.get('id_metodo_pago')
            imagen_comprobante = request.FILES.get('comprobanteahorro')
            
            try:
                monto = float(monto_ahorro)
                metodo = metodos_activos.filter(idmetodopago=id_metodo).first()
                
                # Asumimos el primer bingo disponible temporalmente si tu modelo exige idbingo
                bingo_vinculado = Bingo.objects.exclude(estadobingo='Finalizado').first()
                
                if metodo and monto > 0 and imagen_comprobante:
                    Ahorro.objects.create(
                        idsocio=socio,
                        idbingo=bingo_vinculado, # Requerido por tu modelo
                        idmetodopago=metodo,
                        tipoahorro='Voluntario',
                        montoahorro=monto,
                        comprobanteahorro=imagen_comprobante,
                        fechaahorro=timezone.now(),
                        estadoahorro='Pendiente' # Estado inicial hasta que admin lo valide
                    )
                    messages.success(request, "Tu depósito de ahorro ha sido reportado. Espera la confirmación del administrador.")
                else:
                    messages.error(request, "Datos inválidos o falta adjuntar el comprobante.")
            except ValueError:
                messages.error(request, "El monto ingresado no es válido.")
                
            return redirect('ahorro')

        # 2. LOGICA PARA SOLICITAR RETIROS (Ya la tenías)
        elif action == 'solicitar_retiro':
            monto_retiro = request.POST.get('monto_retiro')
            try:
                monto_retiro = float(monto_retiro)
                if 0 < monto_retiro <= float(total_ahorrado):
                    Ahorro.objects.create(
                        idsocio=socio,
                        idbingo=Bingo.objects.first(), # Requerido
                        tipoahorro='Voluntario',
                        montoahorro=-monto_retiro, 
                        fechaahorro=timezone.now(),
                        estadoahorro='Pendiente' # Adaptamos el estado
                    )
                    messages.success(request, f"Solicitud de retiro de ${monto_retiro} enviada a administración para su aprobación.")
                else:
                    messages.error(request, "Monto inválido o fondos insuficientes en tu libreta.")
            except ValueError:
                messages.error(request, "Por favor ingresa un monto numérico válido.")
            return redirect('ahorro')

    contexto = {
        'socio': socio,
        'historial_ahorros': historial_ahorros,
        'total_ahorrado': total_ahorrado,
        'metodos_activos': metodos_activos # Enviamos las cuentas al template
    }
    return render(request, 'cuentas/ahorro.html', contexto)

def regalo(request): return render(request, 'cuentas/regalo.html')
@login_required
def control_aportes(request):
    
    if request.method == 'POST':
        if not request.user.is_staff:
            messages.error(request, "Solo los administradores pueden registrar aportes.")
            return redirect('inicio')

        id_socio = request.POST.get('id_socio')
        id_bingo = request.POST.get('id_bingo')
        numero_semana = request.POST.get('numero_semana')
        monto = request.POST.get('monto')
        
        try:
            socio = get_object_or_404(Socio, idsocio=id_socio)
            bingo = get_object_or_404(Bingo, idbingo=id_bingo)
            
            # Buscar si el aporte ya existe (para actualizarlo) o crearlo nuevo
            aporte, creado = AporteSemanal.objects.get_or_create(
                idsocio=socio,
                idbingo=bingo,
                numerosemana=numero_semana,
                defaults={
                    'montoaporte': monto,
                    'estadoaporte': 'Al Dia',
                    'fechaplanificadadada': timezone.now() # Fecha de registro
                }
            )
            
            # Si ya existía, simplemente lo marcamos como pagado
            if not creado:
                aporte.montoaporte = monto
                aporte.estadoaporte = 'Al Dia'
                aporte.fechaplanificadadada = timezone.now()
                aporte.save()

            messages.success(request, f"Aporte de la semana {numero_semana} registrado exitosamente para el socio {socio.primernombresocio}.")
            
            # Redirigir a la misma vista manteniendo el bingo seleccionado
            return redirect(f"/control_aportes/?bingo_id={id_bingo}")
            
        except Exception as e:
            messages.error(request, f"Error al registrar el aporte: {str(e)}")
            return redirect('control_aportes')

    # =========================================================
    # LÓGICA EXISTENTE: MOSTRAR LA MATRIZ (PETICIONES GET)
    # =========================================================
    id_bingo = request.GET.get('bingo_id')
    if id_bingo:
        bingo_seleccionado = Bingo.objects.filter(idbingo=id_bingo).first()
    else:
        bingo_seleccionado = Bingo.objects.filter(estadobingo='En Curso').first() or Bingo.objects.order_by('-fechaprogramadabingo').first()

    if not bingo_seleccionado:
        return render(request, 'administrador/control_aportes.html', {'error': 'No hay eventos de bingo creados.'})

    socios = Socio.objects.filter(estadosocio='Activo').order_by('primerapellidosocio', 'primernombresocio')
    aportes = AporteSemanal.objects.filter(idbingo=bingo_seleccionado).select_related('idsocio')

    semanas_query = aportes.values_list('numerosemana', flat=True).distinct().order_by('numerosemana')
    semanas = list(semanas_query) if semanas_query.exists() else list(range(1, 6))

    matriz_socios = {}
    for socio in socios:
        matriz_socios[socio.idsocio] = {
            'objeto_socio': socio,
            'semanas_data': {sem: None for sem in semanas},
            'total_acumulado': Decimal('0.00'),
            'tiene_atrasos': False
        }

    for aporte in aportes:
        s_id = aporte.idsocio_id
        if s_id in matriz_socios:
            if aporte.numerosemana in matriz_socios[s_id]['semanas_data']:
                matriz_socios[s_id]['semanas_data'][aporte.numerosemana] = {
                    'monto': aporte.montoaporte,
                    'estado': aporte.estadoaporte,
                    'id_aporte': aporte.idaporte
                }
                if aporte.estadoaporte == 'Al Dia':
                    matriz_socios[s_id]['total_acumulado'] += aporte.montoaporte
                elif aporte.estadoaporte == 'Atrasado':
                    matriz_socios[s_id]['tiene_atrasos'] = True

    context = {
        'bingo_seleccionado': bingo_seleccionado,
        'todos_los_bingos': Bingo.objects.all().order_by('-fechaprogramadabingo'),
        'semanas': semanas,
        'matriz_socios': matriz_socios.values(),
    }
    return render(request, 'administrador/control_aportes.html', context)
@login_required
def creditos(request):
    """
    Vista para que el Socio vea sus préstamos activos y solicite nuevos créditos.
    """
    socio = Socio.objects.filter(cisocio=request.user.username).first()
    if not socio or socio.estadosocio != 'Activo':
        messages.error(request, "Acceso denegado: Solo los socios activos pueden solicitar créditos.")
        return redirect('inicio')

    mis_prestamos = Prestamo.objects.filter(idsocio=socio).order_by('-fechasolicitud')
    
    # NUEVO: Obtenemos a los otros socios para que salgan en la lista de Garantes (excluyendo al socio actual)
    lista_socios = Socio.objects.filter(estadosocio='Activo').exclude(idsocio=socio.idsocio)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'solicitar_prestamo' and socio:
            monto_solicitado = request.POST.get('montoprestamosolicitado')
            numerocuotas = request.POST.get('numerocuotas')
            tasainteres = request.POST.get('tasainteres')
            fechavencimiento = request.POST.get('fechavencimiento')
            idgarante1 = request.POST.get('idgarante1')
            idgarante2 = request.POST.get('idgarante2')
            
            try:
                monto = float(monto_solicitado)
                cuotas = int(numerocuotas)
                tasa = Decimal(tasainteres)
                
                if monto > 0 and cuotas > 0 and tasa >= 0:
                    interes_calculado = (Decimal(str(monto)) * tasa) / Decimal('100')
                    monto_total = Decimal(str(monto)) + interes_calculado
                    
                    garante1_obj = Socio.objects.filter(idsocio=idgarante1).first() if idgarante1 else None
                    garante2_obj = Socio.objects.filter(idsocio=idgarante2).first() if idgarante2 else None
                    
                    Prestamo.objects.create(
                        idsocio=socio,
                        idgarante1=garante1_obj,
                        idgarante2=garante2_obj,
                        montoprestamosolicitado=monto,
                        tasainteres=tasa,
                        numerocuotas=cuotas,
                        montototalpagar=monto_total,
                        saldopendiente=monto_total,
                        fechasolicitud=timezone.now(),
                        fechavencimiento=fechavencimiento, 
                        estadoprestamo='Solicitado'
                    )
                    messages.success(request, f"¡Tu solicitud de crédito por ${monto_solicitado} ha sido enviada con éxito!")
                else:
                    messages.error(request, "El monto y las cuotas deben ser valores mayores a 0.")
            except Exception as e:
                messages.error(request, f"Ocurrió un error al procesar tu solicitud: {str(e)}")
                
            return redirect('perfil')

    contexto = {
        'socio': socio,
        'mis_prestamos': Prestamo.objects.filter(idsocio=socio),
        'lista_socios': Socio.objects.filter(estadosocio='Activo').exclude(idsocio=socio.idsocio)
    }
    return render(request, 'negocio/creditos.html', contexto)

@login_required
def metodos_pago(request):
    """
    Vista informativa para que el Socio vea los métodos de pago aceptados por la cooperativa.
    (Ej. Números de cuenta donde puede realizar depósitos).
    """
    if not request.user.is_authenticated:
        return redirect('inicio_sesion')
        
    metodos_activos = MetodoPago.objects.filter(estadometodopago=True)
    
    contexto = {
        'metodos_activos': metodos_activos
    }
    return render(request, 'negocio/metodos_pago.html', contexto)

@login_required
def pago(request):
    """
    Vista para que el Socio reporte o registre un pago realizado a sus préstamos.
    """
    socio = Socio.objects.filter(cisocio=request.user.username).first()
    if not socio:
        messages.warning(request, "Debes ser Socio para registrar pagos de préstamos.")
        return redirect('inicio')

    # Préstamos que aún tienen saldo pendiente
    prestamos_activos = Prestamo.objects.filter(idsocio=socio).exclude(estadoprestamo='Liquidado')
    metodos_activos = MetodoPago.objects.filter(estadometodopago='Activo')
    historial_pagos = Pago.objects.filter(idprestamo__idsocio=socio).order_by('-fechapago')

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'registrar_pago':
            # CORRECCIÓN 1: Nombres exactos de los campos del formulario HTML
            id_prestamo = request.POST.get('id_prestamo')
            id_metodo = request.POST.get('id_metodo_pago') 
            monto_pagado = request.POST.get('monto_pagado')
            
            # CORRECCIÓN 2: Capturamos la imagen del comprobante
            imagen_comprobante = request.FILES.get('imagencomprobante')

            try:
                monto = float(monto_pagado)
                # Validar que el préstamo pertenece al socio y existe
                prestamo = prestamos_activos.filter(idprestamo=id_prestamo).first()
                metodo = metodos_activos.filter(idmetodopago=id_metodo).first()

                # CORRECCIÓN 3: Validamos que exista la imagen (es requerida por tu modelo)
                if prestamo and metodo and monto > 0:
                    if imagen_comprobante:
                        Pago.objects.create(
                            idprestamo=prestamo,
                            idmetodopago=metodo,
                            montopagado=monto, # Coincide con el modelo
                            comprobantepago=imagen_comprobante, # Se guarda la imagen
                            fechapago=timezone.now(),
                            estadopago='Pendiente' # Queda pendiente de verificación
                        )
                        messages.success(request, "Tu comprobante de pago ha sido registrado. Un administrador verificará la transacción pronto.")
                    else:
                        messages.error(request, "Es obligatorio subir una foto o captura del comprobante de pago.")
                else:
                    messages.error(request, "Datos inválidos. Por favor, verifica el préstamo y el método de pago seleccionado.")
            except (ValueError, TypeError):
                messages.error(request, "El monto ingresado no es válido.")
        
        return redirect('pago')

    contexto = {
        'socio': socio,
        'prestamos_activos': prestamos_activos,
        'metodos_activos': metodos_activos,
        'historial_pagos': historial_pagos
    }
    return render(request, 'negocio/pago.html', contexto)

# IMPORTANTE: Eliminamos el @login_required de aquí arriba
def sala_espera(request, id_partida):
    # 1. EL GUARDIA: Interceptor de invitados
    if not request.user.is_authenticated:
        # Si no tiene cuenta, lo mandamos a tu nuevo pantallazo de bloqueo
        return render(request, 'cuentas/acceso_denegado.html')

    # 2. LÓGICA NORMAL: El resto de tu código queda intacto
    jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
    if not jugador:
        return redirect('registro_jugador')

    # Mantengo tu búsqueda exacta con idpartidabingo
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    # Si la partida ya empezó, los metemos directo al tablero
    if partida.estadopartida == 'En Juego':
        return redirect('tablero_tiempo_real', id_partida=partida.idpartidabingo)
    
    # =========================================================
    # EL PORTERO VIP: REDIRECCIÓN ABSOLUTA DESDE EL SERVIDOR
    # =========================================================
    if partida.estadopartida in ['Verificando', 'Desempate'] and partida.idbingadores:
        ids_vip = [int(i.strip()) for i in str(partida.idbingadores).split(',') if i.strip()]
        if jugador.idjugador in ids_vip:
            return redirect('sala_espera_desempate', id_partida=partida.idpartidabingo)
    # =========================================================
    
    # >>> INYECTA TODO ESTE BLOQUE NUEVO AQUÍ <<<
    # =========================================================
    # FIX FASE 1: REGISTRO DE SESIÓN PARA LA NUEVA RONDA
    # =========================================================
    try:
        plataforma, _ = PlataformaJuego.objects.get_or_create(
            nombreplataforma='Web Oficial',
            defaults={'urlplataforma': request.build_absolute_uri('/'), 'estadoplataforma': True}
        )
        
        user_agent = request.META.get('HTTP_USER_AGENT', 'Desconocido')
        dispositivo = 'Dispositivo Móvil' if any(x in user_agent for x in ['Mobile', 'Android', 'iPhone']) else 'Tablet' if 'iPad' in user_agent else 'PC / Escritorio'
        navegador = 'Google Chrome' if 'Chrome' in user_agent else 'Apple Safari' if 'Safari' in user_agent else 'Mozilla Firefox' if 'Firefox' in user_agent else 'Otro Navegador'

        with transaction.atomic():
            # 1. Cerramos CUALQUIER sesión activa anterior del jugador (lo sacamos de la ronda vieja)
            SesionJuego.objects.filter(
                idjugador=jugador, 
                estadosesion='Activa'
            ).update(
                estadosesion='Finalizada', 
                fechafinsesion=timezone.now(), 
                motivocierre='Traslado automático a nueva ronda'
            )

            # 2. Creamos la nueva sesión oficial para ESTA ronda de espera
            SesionJuego.objects.create(
                idplataforma=plataforma,
                idjugador=jugador,
                idpartida=partida,
                fechainiciosesion=timezone.now(),
                ipconexion=obtener_ip_cliente(request),
                dispositivoconexion=dispositivo,
                estadosesion='Activa',
                navegadorweb=navegador,
                tokenconexion=str(uuid.uuid4())
            )
    except Exception as e:
        print(f"Error al registrar la sesión en sala de espera: {str(e)}")
    # =========================================================
    
    # =========================================================
    # FIX ANTI-FANTASMAS: SOLO JUGADORES CON CONEXIÓN ACTIVA
    # =========================================================
    jugadores_en_sala = Jugador.objects.filter(
        sesionjuego__idpartida=partida,
        sesionjuego__estadosesion='Activa'
    ).distinct().order_by('aliasjugador')
    # =========================================================
        
    mensajes_historial = MensajeChat.objects.filter(idbingo=partida.idbingo).order_by('fechahora')
    
    contexto = {
        'partida': partida,
        'jugador': jugador,
        'jugadores_en_sala': jugadores_en_sala,
        'mensajes_historial': mensajes_historial # Añadir esta línea
    }
    return render(request, 'partida/sala_espera.html', contexto)

@login_required
def sala_espera_desempate(request, id_partida):
    # Modificada para recibir el ID dinámico de la partida
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    # Si el administrador ya resolvió y finalizó, no tiene sentido que esperen aquí
    if partida.estadopartida == 'Finalizada':
        return redirect('inicio')
    
    # =========================================================
    # FIX ANTI-FANTASMAS: SOLO JUGADORES CON CONEXIÓN ACTIVA
    # =========================================================
    jugadores_en_sala = Jugador.objects.filter(
        sesionjuego__idpartida=partida,
        sesionjuego__estadosesion='Activa'
    ).distinct().order_by('aliasjugador')
    # =========================================================

    jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
    mensajes_historial = MensajeChat.objects.filter(idbingo=partida.idbingo).order_by('fechahora')

    contexto = {
        'partida': partida,
        'jugador': jugador,
        'jugadores_en_sala': jugadores_en_sala,
        'mensajes_historial': mensajes_historial
    }
    return render(request, 'partida/sala_espera_desempate.html', contexto)

def estado_partida_json(request, id_partida):
    """
    Endpoint API público/interno para que los jugadores consulten el estado 
    de la ronda en tiempo real sin recargar la página.
    """
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    return JsonResponse({
        'estado': partida.estadopartida,
        'hay_desempate': partida.haydesempate,
        'ganador': partida.idjugadororganador.aliasjugador if partida.idjugadororganador else None,
        'premio_efectivo': str(partida.valorpremio)
    })

@login_required 
def tablero_tiempo_real(request, id_partida):
    jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
    if not jugador:
        messages.warning(request, "Necesitas un perfil de jugador para entrar a la sala.")
        return redirect('inicio')

    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)

    # 1. Seguridad: Redirección automática si el admin pausa o el juego termina
    if partida.estadopartida in ['Verificando', 'Desempate']:
        return redirect('sala_espera_desempate', id_partida=partida.idpartidabingo)
    elif partida.estadopartida == 'Finalizada':
        messages.info(request, "Esta ronda ha finalizado.")
        return redirect('inicio')

    # =========================================================
    # DETECTOR AUTOMÁTICO Y SILENCIOSO DE SESIÓN DE JUEGO (LOG)
    # =========================================================
    try:
        # A. Obtener o crear la Plataforma Base para la Web del sistema
        plataforma, _ = PlataformaJuego.objects.get_or_create(
            nombreplataforma='Web Oficial',
            defaults={
                'urlplataforma': request.build_absolute_uri('/'),
                'descripcionplataforma': 'Acceso nativo desde la aplicación web.',
                'estadoplataforma': True
            }
        )
        
        # B. Extraer metadatos del Navegador y Dispositivo
        user_agent = request.META.get('HTTP_USER_AGENT', 'Desconocido')
        
        dispositivo = 'PC / Escritorio'
        if 'Mobile' in user_agent or 'Android' in user_agent or 'iPhone' in user_agent:
            dispositivo = 'Dispositivo Móvil'
        elif 'iPad' in user_agent or 'Tablet' in user_agent:
            dispositivo = 'Tablet'

        navegador = 'Otro Navegador'
        if 'Chrome' in user_agent: navegador = 'Google Chrome'
        elif 'Safari' in user_agent and 'Chrome' not in user_agent: navegador = 'Apple Safari'
        elif 'Firefox' in user_agent: navegador = 'Mozilla Firefox'
        elif 'Edge' in user_agent: navegador = 'Microsoft Edge'

        # C. Control de Concurrencia (Anti-Clonación) en una transacción aislada
        with transaction.atomic():
            # Si el jugador ya tenía otra pestaña u otro dispositivo activo en esta ronda, lo finalizamos
            SesionJuego.objects.filter(
                idjugador=jugador, 
                idpartida=partida, 
                estadosesion='Activa'
            ).update(
                estadosesion='Finalizada', 
                fechafinsesion=timezone.now(), 
                motivocierre='Nueva conexión establecida'
            )

            # Creamos la nueva sesión activa oficial para este dispositivo
            SesionJuego.objects.create(
                idplataforma=plataforma,
                idjugador=jugador,
                idpartida=partida,
                fechainiciosesion=timezone.now(),
                ipconexion=obtener_ip_cliente(request),
                dispositivoconexion=dispositivo,
                estadosesion='Activa',
                navegadorweb=navegador,
                tokenconexion=str(uuid.uuid4())
            )
    except Exception as e:
        # Fallback de seguridad: Si falla el log de sesión, no bloqueamos el juego, solo lo dejamos pasar
        print(f"Error silencioso al registrar la sesión de auditoría: {str(e)}")
    # =========================================================

    # 2. Traer los cartones que ESTE jugador compró para ESTA partida
    cartones_asignados = CartonPartidaBingo.objects.filter(
        idjugador=jugador,
        idpartida=partida
    ).select_related('idcarton')

    # 3. Procesar las bolas cantadas para saber cuáles colorear
    bolas_str = partida.bolascantadas.replace('B','').replace('I','').replace('N','').replace('G','').replace('O','')
    bolas_llamadas = [int(b.strip()) for b in bolas_str.split(',') if b.strip().isdigit()]

    # 4. Preparar las matrices para que HTML las dibuje fácilmente (fila por fila)
    for asignacion in cartones_asignados:
        matriz = asignacion.idcarton.matriznumeros
        if isinstance(matriz, str):
            matriz = json.loads(matriz.replace("'", '"'))
        
        filas = []
        for i in range(5):
            fila = [
                matriz['B'][i], matriz['I'][i], matriz['N'][i], matriz['G'][i], matriz['O'][i]
            ]
            filas.append(fila)
        asignacion.filas_matriz = filas

    # 5. Obtener la lista de todos los jugadores únicos en esta partida
    jugadores_en_sala = Jugador.objects.filter(
        sesionjuego__idpartida=partida,
        sesionjuego__estadosesion='Activa'
    ).distinct().order_by('aliasjugador')

    mensajes_historial = MensajeChat.objects.filter(idbingo=partida.idbingo).order_by('fechahora')

    contexto = {
        'partida': partida,
        'jugador': jugador,
        'cartones_asignados': cartones_asignados,
        'bolas_llamadas': bolas_llamadas,
        'jugadores_en_sala': jugadores_en_sala,
        'mensajes_historial': mensajes_historial, # Añadir esta línea
    }
    return render(request, 'partida/tablero_tiempo_real.html', contexto)

def obtener_ip_cliente(request):
    """Función helper para extraer la dirección IP real del jugador"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

@login_required
def sesion_juego(request, id_partida):
    jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
    if not jugador:
        return redirect('registro_jugador')
        
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    # 1. Obtener o crear la Plataforma Base para la Web del sistema
    plataforma, _ = PlataformaJuego.objects.get_or_create(
        nombreplataforma='Web Oficial',
        defaults={
            'urlplataforma': request.build_absolute_uri('/'),
            'descripcionplataforma': 'Acceso nativo desde la aplicación web.',
            'estadoplataforma': True
        }
    )
    
    # 2. Extraer metadatos del Navegador y Dispositivo
    user_agent = request.META.get('HTTP_USER_AGENT', 'Desconocido')
    
    # Análisis simple del dispositivo
    dispositivo = 'PC / Escritorio'
    if 'Mobile' in user_agent or 'Android' in user_agent or 'iPhone' in user_agent:
        dispositivo = 'Dispositivo Móvil'
    elif 'iPad' in user_agent or 'Tablet' in user_agent:
        dispositivo = 'Tablet'

    # Limpieza simple del nombre del navegador
    navegador = 'Otro Navegador'
    if 'Chrome' in user_agent: navegador = 'Google Chrome'
    elif 'Safari' in user_agent and 'Chrome' not in user_agent: navegador = 'Apple Safari'
    elif 'Firefox' in user_agent: navegador = 'Mozilla Firefox'
    elif 'Edge' in user_agent: navegador = 'Microsoft Edge'

    # 3. Registrar la Sesión de Juego en la Base de Datos
    # Usamos transacciones seguras para evitar duplicados críticos
    with transaction.atomic():
        # Cerramos posibles sesiones previas 'Activas' de este jugador en esta ronda
        SesionJuego.objects.filter(
            idjugador=jugador, 
            idpartida=partida, 
            estadosesion='Activa'
        ).update(estadosesion='Finalizada', fechafinsesion=timezone.now(), motivocierre='Nueva conexión establecida')

        # Creamos la nueva sesión oficial
        sesion = SesionJuego.objects.create(
            idplataforma=plataforma,
            idjugador=jugador,
            idpartida=partida,
            fechainiciosesion=timezone.now(),
            ipconexion=obtener_ip_cliente(request),
            dispositivoconexion=dispositivo,
            estadosesion='Activa',
            navegadorweb=navegador,
            tokenconexion=str(uuid.uuid4()) # Token criptográfico único
        )

    contexto = {
        'partida': partida,
        'sesion': sesion
    }
    return render(request, 'partida/sesion_juego.html', contexto)

@login_required
def tablero_admin(request, id_partida):
    if not request.user.is_staff:
        messages.error(request, "Acceso denegado. Zona exclusiva de administración.")
        return redirect('inicio')
        
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    # =======================================================
    # NUEVO: BOTÓN MANUAL DE INICIO DE PARTIDA Y FIX DE ESTADO
    # =======================================================
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'iniciar_partida' and partida.estadopartida == 'Programada':
            # 1. Actualizamos la ronda actual
            partida.estadopartida = 'En Juego'
            partida.horainiciopartida = timezone.now()
            partida.save()
            
            # 2. FIX: ACTUALIZAR EL BINGO PADRE A 'EN CURSO'
            bingo_padre = partida.idbingo
            if bingo_padre.estadobingo == 'Programado':
                bingo_padre.estadobingo = 'En Curso'
                bingo_padre.save()
            
            # 3. El Árbitro sopla el silbato por WebSockets
            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f'bingo_partida_{partida.idpartidabingo}',
                {'type': 'evento_partida', 'datos': {'evento': 'estado_cambiado', 'nuevo_estado': 'En Juego'}}
            )
            messages.success(request, f"¡Pitazo inicial! La ronda ha comenzado y el Bingo está En Curso.")
            return redirect('tablero_admin', id_partida=partida.idpartidabingo)
    # =======================================================

    # 1. Limpiar y obtener las bolas cantadas como una lista de enteros
    bolas_str = partida.bolascantadas.replace('B','').replace('I','').replace('N','').replace('G','').replace('O','')
    bolas_llamadas = [int(b.strip()) for b in bolas_str.split(',') if b.strip().isdigit()]

    # 2. Construir la estructura del Tablero Maestro (1 al 75)
    tablero_maestro = {
        'B': {'rango': range(1, 16), 'color': 'primary'},     
        'I': {'rango': range(16, 31), 'color': 'danger'},     
        'N': {'rango': range(31, 46), 'color': 'secondary'},  
        'G': {'rango': range(46, 61), 'color': 'success'},    
        'O': {'rango': range(61, 76), 'color': 'warning'}     
    }

    # NUEVO: Obtener la lista de todos los jugadores únicos en esta partida para el radar
    jugadores_en_sala = Jugador.objects.filter(
        sesionjuego__idpartida=partida,
        sesionjuego__estadosesion='Activa'
    ).distinct().order_by('aliasjugador')

    contexto = {
        'partida': partida,
        'bolas_llamadas': bolas_llamadas,
        'tablero_maestro': tablero_maestro,
        'jugadores_en_sala': jugadores_en_sala, # <--- Enviamos los jugadores al Tablero Admin
    }
    return render(request, 'partida/tablero_admin.html', contexto)

@login_required
def desempate_admin(request, id_partida):
    if not request.user.is_staff:
        return redirect('inicio')
        
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    channel_layer = get_channel_layer() 
    
    if partida.estadopartida == 'En Juego':
        partida.estadopartida = 'Verificando'
        partida.save()
        
        async_to_sync(channel_layer.group_send)(
            f'bingo_partida_{id_partida}',
            {'type': 'evento_partida', 'datos': {'evento': 'estado_cambiado', 'nuevo_estado': 'Verificando'}}
        )

    if request.method == 'POST':
        decision = request.POST.get('decision_desempate')
        
        if decision == 'si':
            partida.estadopartida = 'Desempate'
            partida.haydesempate = True
            partida.save()
            
            async_to_sync(channel_layer.group_send)(
                f'bingo_partida_{id_partida}',
                {'type': 'evento_partida', 'datos': {'evento': 'estado_cambiado', 'nuevo_estado': 'Desempate'}}
            )
            
            messages.info(request, "Modo Desempate Activado. Prepare la consola.")
            return redirect('consola_juego', id_partida=partida.idpartidabingo)
            
        elif decision == 'no':
            codigo_ganador = request.POST.get('codigo_ganador_unico')
            resultado = validar_carton_hibrido(codigo_ganador, partida.idpartidabingo)
            
            if resultado['existe'] and resultado['valido']:
                partida.estadopartida = 'Finalizada'
                partida.idjugadororganador_id = resultado['id_jugador']
                partida.horafin = timezone.now() 
                partida.save()
                
                # ==========================================
                # MAGIA FINANCIERA: PAGO AUTOMÁTICO DE PREMIOS
                # ==========================================
                es_pozo_mayor = (partida.premiomaterial == '[POZO_MAYOR]')
                monto_a_pagar = partida.idbingo.premiomayor if es_pozo_mayor else partida.valorpremio
                
                if monto_a_pagar and monto_a_pagar > 0:
                    jugador_ganador = Jugador.objects.get(idjugador=resultado['id_jugador'])
                    tipo_moneda = partida.idbingo.idunidadmonetaria.tipomoneda
                    if tipo_moneda == 'Efectivo':
                        jugador_ganador.saldocreditojugador += monto_a_pagar
                    else:
                        jugador_ganador.saldovirtualjugador += monto_a_pagar
                    jugador_ganador.save()
                
                # Logística del Premio Físico (Si NO es el pozo mayor)
                if not es_pozo_mayor and partida.premiomaterial and partida.premiomaterial != 'Ninguno':
                    partida.estadopremiomaterial = 'Pendiente'
                    
                partida.save()

                # ==========================================
                # ÁRBITRO DIGITAL: RELEVO Y ENRUTAMIENTO (FASE 3)
                # ==========================================
                siguiente_partida = PartidaBingo.objects.filter(
                    idbingo=partida.idbingo,
                    idpartidabingo__gt=partida.idpartidabingo
                ).order_by('idpartidabingo').first()

                if siguiente_partida:
                    # FIX: Eliminamos el cambio de estado a 'En Juego'. 
                    # Ahora la partida se queda 'Programada' y el admin viaja al tablero a iniciarla manualmente.
                    destino_admin = redirect('tablero_admin', id_partida=siguiente_partida.idpartidabingo)
                else:
                    bingo_actual = partida.idbingo
                    bingo_actual.estadobingo = 'Finalizado'
                    bingo_actual.save()
                    # Si ya no hay rondas, lo mandamos al dashboard
                    destino_admin = redirect('dashboard')
                # ==========================================
                
                id_siguiente = siguiente_partida.idpartidabingo if siguiente_partida else None
                
                async_to_sync(channel_layer.group_send)(
                    f'bingo_partida_{id_partida}',
                    {'type': 'evento_partida', 'datos': {
                        'evento': 'estado_cambiado', 
                        'nuevo_estado': 'Finalizada',
                        'ganador': resultado['jugador'],
                        'id_siguiente_partida': id_siguiente # <-- LA LLAVE MÁGICA
                    }}
                )
                
                messages.success(request, f"¡Partida finalizada! Ganador único asignado: {resultado['jugador']}")
                
                # ¡Vuelo directo al nuevo destino!
                return destino_admin
            else:
                messages.error(request, "El código ingresado no es válido o no completó el cartón.")
                return redirect('desempate_admin', id_partida=partida.idpartidabingo)

    # =========================================================
    # ESCÁNER DE GANADORES WEB EN TIEMPO REAL (AQUÍ AFUERA DEL POST)
    # =========================================================
    import json
    bolas_str = partida.bolascantadas.replace('B','').replace('I','').replace('N','').replace('G','').replace('O','')
    bolas_llamadas = [int(b.strip()) for b in bolas_str.split(',') if b.strip().isdigit()]
    
    patrones = {
        'Tabla Llena': [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24],
        'Las Cuatro Esquinas': [0, 4, 20, 24],
        'En Diagonal': [0, 6, 12, 18, 24],
        'Forma de X': [0, 4, 6, 8, 12, 16, 18, 20, 24],
        'Forma de Cruz': [2, 7, 10, 11, 12, 13, 14, 17, 22],
        'Marco de Foto': [0,1,2,3,4, 5,9, 10,14, 15,19, 20,21,22,23,24],
        'Linea Vertical': [2, 7, 12, 17, 22],
        'Forma de L': [0, 5, 10, 15, 20, 21, 22, 23, 24],
        'Forma de C': [0,1,2,3,4, 5, 10, 15, 20,21,22,23,24],
        'Forma de T': [0,1,2,3,4, 7, 12, 17, 22],
        'Forma de U': [0,4, 5,9, 10,14, 15,19, 20,21,22,23,24],
        'Forma de H': [0,4, 5,9, 10,11,12,13,14, 15,19, 20,24],
        'Forma de Z': [0,1,2,3,4, 8, 12, 16, 20,21,22,23,24],
        'Forma de Flecha': [2, 6, 8, 12, 17, 22]
    }
    marcadas_requeridas = patrones.get(partida.modalidad_victoria, patrones['Tabla Llena'])
    
    cartones_en_juego = CartonPartidaBingo.objects.filter(idpartida=partida).select_related('idcarton', 'idjugador')
    ganadores_web = []
    
    for c in cartones_en_juego:
        matriz = c.idcarton.matriznumeros
        if isinstance(matriz, str):
            try: matriz = json.loads(matriz.replace("'", '"'))
            except: continue
            
        celdas = []
        for i in range(5):
            celdas.extend([matriz['B'][i], matriz['I'][i], matriz['N'][i], matriz['G'][i], matriz['O'][i]])
            
        es_ganador = True
        for idx in marcadas_requeridas:
            if idx == 12: continue # La celda del medio (FREE)
            if int(celdas[idx]) not in bolas_llamadas:
                es_ganador = False
                break
                
        if es_ganador:
            ganadores_web.append(c)
    # =========================================================

    contexto = {
        'partida': partida,
        'ganadores_web': ganadores_web
    }
    return render(request, 'partida/desempate_admin.html', contexto)

@login_required
def consola_juego(request, id_partida):
    if not request.user.is_staff:
        return redirect('inicio')
        
    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    if partida.estadopartida == 'Finalizada':
        messages.info(request, "Esta partida ya ha finalizado.")
        return redirect('dashboard')
    
    candidatos_ids = []
    if partida.idbingadores:
        candidatos_ids = [int(id_str) for id_str in partida.idbingadores.split(',') if id_str.strip()]
        
    candidatos = Jugador.objects.filter(idjugador__in=candidatos_ids)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'agregar_candidato':
            codigo = request.POST.get('codigo_carton')
            resultado = validar_carton_hibrido(codigo, partida.idpartidabingo)
            
            if resultado['existe'] and resultado['valido']:
                nuevo_id = str(resultado['id_jugador'])
                
                ids_actuales = partida.idbingadores.split(',') if partida.idbingadores else []
                if nuevo_id not in ids_actuales:
                    if partida.idbingadores:
                        partida.idbingadores += f",{nuevo_id}"
                    else:
                        partida.idbingadores = nuevo_id
                    partida.save()
                    
                    # ==========================================
                    # EL BOLETO VIP: TELETRANSPORTE AL JUGADOR
                    # ==========================================
                    channel_layer = get_channel_layer()
                    async_to_sync(channel_layer.group_send)(
                        f'bingo_partida_{id_partida}',
                        {
                            'type': 'evento_partida',
                            'datos': {
                                'evento': 'invitacion_vip',
                                'id_jugador': str(resultado['id_jugador']) # <--- EL CAMBIO CLAVE AQUÍ
                            }
                        }
                    )
                    # ==========================================
                    
                    messages.success(request, f"¡Cartón verificado! {resultado['jugador']} agregado al desempate.")
                else:
                    messages.warning(request, "Este jugador ya está en la lista de desempate.")
            else:
                messages.error(request, "Código inválido o cartón incompleto.")
            
            return redirect('consola_juego', id_partida=id_partida)
        
        # =========================================================
        # NUEVO: MOTOR DE MEMORIA Y WEBSOCKET (JUEZ SUPREMO)
        # =========================================================
        elif action == 'registrar_tiro_desempate':
            id_jugador_tiro = request.POST.get('id_jugador_tiro')
            numero_tiro = int(request.POST.get('numero_tiro'))
            
            # 1. Guardamos el tiro en la Bóveda JSON
            sorteo = partida.sorteodesempate or {}
            sorteo[str(id_jugador_tiro)] = numero_tiro
            partida.sorteodesempate = sorteo
            partida.save()
            
            # 2. Verificamos si TODOS los candidatos ya tiraron
            ids_actuales = [str(i.strip()) for i in str(partida.idbingadores).split(',') if i.strip()]
            completado = all(candidato in sorteo for candidato in ids_actuales)
            
            if completado:
                # 3. Django decide quién ganó (matemáticamente)
                ganador_id = max(sorteo, key=sorteo.get)
                ganador_numero = sorteo[ganador_id]
                ganador_obj = Jugador.objects.filter(idjugador=int(ganador_id)).first()
                ganador_nombre = ganador_obj.aliasjugador if ganador_obj else "Jugador Oficial"
                
                # 4. Disparamos el misil por WebSocket a la consola
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    f'bingo_partida_{id_partida}',
                    {
                        'type': 'evento_partida',
                        'datos': {
                            'evento': 'desempate_completado',
                            'ganador_id': ganador_id,
                            'ganador_numero': ganador_numero,
                            'ganador_nombre': ganador_nombre
                        }
                    }
                )
            
            return JsonResponse({'status': 'ok', 'completado': completado})
        # =========================================================
            
        elif action == 'resolver_desempate':
            ganador_id = request.POST.get('ganador_final')
            bola_mayor = request.POST.get('bola_mayor')
            
            if ganador_id and bola_mayor:
                partida.idjugadororganador_id = ganador_id
                partida.bolamayordesempate = bola_mayor
                partida.estadopartida = 'Finalizada'
                partida.horafin = timezone.now() 
                partida.save()
                
                # ==========================================
                # MAGIA FINANCIERA: PAGO AUTOMÁTICO DE PREMIOS
                # ==========================================
                es_pozo_mayor = (partida.premiomaterial == '[POZO_MAYOR]')
                monto_a_pagar = partida.idbingo.premiomayor if es_pozo_mayor else partida.valorpremio
                
                if monto_a_pagar and monto_a_pagar > 0:
                    jugador_ganador = Jugador.objects.get(idjugador=ganador_id)
                    tipo_moneda = partida.idbingo.idunidadmonetaria.tipomoneda
                    if tipo_moneda == 'Efectivo':
                        jugador_ganador.saldocreditojugador += monto_a_pagar
                    else:
                        jugador_ganador.saldovirtualjugador += monto_a_pagar
                    jugador_ganador.save()
                    
                # Logística del Premio Físico
                if not es_pozo_mayor and partida.premiomaterial and partida.premiomaterial != 'Ninguno':
                    partida.estadopremiomaterial = 'Pendiente'

                partida.save()
                
                # ==========================================
                # ÁRBITRO DIGITAL: RELEVO Y ENRUTAMIENTO (FASE 3)
                # ==========================================
                siguiente_partida = PartidaBingo.objects.filter(
                    idbingo=partida.idbingo,
                    idpartidabingo__gt=partida.idpartidabingo
                ).order_by('idpartidabingo').first()

                if siguiente_partida:
                    # FIX: Eliminamos el auto-arranque de la partida.
                    # Viaje directo al tablero para inicio manual.
                    destino_admin = redirect('tablero_admin', id_partida=siguiente_partida.idpartidabingo)
                else:
                    bingo_actual = partida.idbingo
                    bingo_actual.estadobingo = 'Finalizado'
                    bingo_actual.save()
                    # Preparamos viaje al dashboard
                    destino_admin = redirect('dashboard')
                # ==========================================
                
                id_siguiente = siguiente_partida.idpartidabingo if siguiente_partida else None
                
                ganador_obj = Jugador.objects.get(idjugador=ganador_id)
                channel_layer = get_channel_layer()
                async_to_sync(channel_layer.group_send)(
                    f'bingo_partida_{id_partida}',
                    {'type': 'evento_partida', 'datos': {
                        'evento': 'estado_cambiado', 
                        'nuevo_estado': 'Finalizada',
                        'ganador': ganador_obj.aliasjugador,
                        'id_siguiente_partida': id_siguiente # <-- LA LLAVE MÁGICA
                    }}
                )
                
                messages.success(request, "¡Desempate resuelto! El ganador ha sido registrado y la ronda ha finalizado.")
                return destino_admin
            else:
                messages.error(request, "Debe seleccionar un ganador e ingresar la bola mayor.")
                return redirect('consola_juego', id_partida=id_partida)

    # =========================================================
    # ESCÁNER DE GANADORES WEB EN TIEMPO REAL (RADAR WEB)
    # =========================================================
    import json
    bolas_str = partida.bolascantadas.replace('B','').replace('I','').replace('N','').replace('G','').replace('O','')
    bolas_llamadas = [int(b.strip()) for b in bolas_str.split(',') if b.strip().isdigit()]
    
    patrones = {
        'Tabla Llena': [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24],
        'Las Cuatro Esquinas': [0, 4, 20, 24],
        'En Diagonal': [0, 6, 12, 18, 24],
        'Forma de X': [0, 4, 6, 8, 12, 16, 18, 20, 24],
        'Forma de Cruz': [2, 7, 10, 11, 12, 13, 14, 17, 22],
        'Marco de Foto': [0,1,2,3,4, 5,9, 10,14, 15,19, 20,21,22,23,24],
        'Linea Vertical': [2, 7, 12, 17, 22],
        'Forma de L': [0, 5, 10, 15, 20, 21, 22, 23, 24],
        'Forma de C': [0,1,2,3,4, 5, 10, 15, 20,21,22,23,24],
        'Forma de T': [0,1,2,3,4, 7, 12, 17, 22],
        'Forma de U': [0,4, 5,9, 10,14, 15,19, 20,21,22,23,24],
        'Forma de H': [0,4, 5,9, 10,11,12,13,14, 15,19, 20,24],
        'Forma de Z': [0,1,2,3,4, 8, 12, 16, 20,21,22,23,24],
        'Forma de Flecha': [2, 6, 8, 12, 17, 22]
    }
    marcadas_requeridas = patrones.get(partida.modalidad_victoria, patrones['Tabla Llena'])
    
    cartones_en_juego = CartonPartidaBingo.objects.filter(idpartida=partida).select_related('idcarton', 'idjugador')
    ganadores_web = []
    
    for c in cartones_en_juego:
        matriz = c.idcarton.matriznumeros
        if isinstance(matriz, str):
            try: matriz = json.loads(matriz.replace("'", '"'))
            except: continue
            
        celdas = []
        for i in range(5):
            celdas.extend([matriz['B'][i], matriz['I'][i], matriz['N'][i], matriz['G'][i], matriz['O'][i]])
            
        es_ganador = True
        for idx in marcadas_requeridas:
            if idx == 12: continue # La celda del medio (FREE)
            if int(celdas[idx]) not in bolas_llamadas:
                es_ganador = False
                break
                
        if es_ganador:
            ganadores_web.append(c)
    # =========================================================

    contexto = {
        'partida': partida,
        'candidatos': candidatos,
        'ganadores_web': ganadores_web # <-- Aquí enviamos a los que bingaron a la plantilla
    }
    return render(request, 'partida/consola_juego.html', contexto)

def inicio_sesion(request):
    if request.user.is_authenticated: return redirect('dashboard' if request.user.is_staff else 'inicio')
    if request.method == 'POST':
        identificador = request.POST.get('identificador')
        password = request.POST.get('password')
        user = User.objects.filter(Q(username=identificador) | Q(email=identificador)).first()
        if user and user.check_password(password):
            if not user.is_active:
                messages.error(request, "Esta cuenta ha sido desactivada o suspendida del sistema.")
                return redirect('inicio_sesion')
            login(request, user)
            socio = Socio.objects.filter(cisocio=user.username).first()
            jugador = Jugador.objects.filter(cedulaidentidadjugador=user.username).first()
            nombre_mostrar = user.first_name
            avatar_url = None
            if jugador:
                nombre_mostrar = jugador.aliasjugador or user.first_name
                if jugador.avatarjugador: avatar_url = jugador.avatarjugador.url
            if socio and not avatar_url:
                if socio.fotosocio: avatar_url = socio.fotosocio.url
            request.session['user_nombre'] = nombre_mostrar
            request.session['avatar_url'] = avatar_url
            messages.success(request, f"¡Bienvenido de vuelta, {nombre_mostrar}!")
            return redirect('dashboard' if user.is_staff else 'inicio')
        else:
            messages.error(request, "Credenciales incorrectas. Verifica tu usuario/cédula/correo y contraseña.")
    return render(request, 'cuentas/inicio_sesion.html')

def cerrar_sesion(request):
    logout(request)
    return redirect('inicio')

def seleccion_registro(request): return render(request, 'cuentas/seleccion_registro.html')

def registro_socio(request):
    if request.method == 'POST':
        primer_nombre, segundo_nombre = request.POST.get('primer_nombre'), request.POST.get('segundo_nombre')
        primer_apellido, segundo_apellido = request.POST.get('primer_apellido'), request.POST.get('segundo_apellido')
        cedula, fecha_nacimiento_str = request.POST.get('cedula'), request.POST.get('fecha_nacimiento')
        telefono_personal, direccion = request.POST.get('telefono_personal'), request.POST.get('direccion')
        sexo, email, password = request.POST.get('sexo'), request.POST.get('email'), request.POST.get('password')

        # ==========================================
        # FIX: ESCUDO DE INTEGRIDAD DE DATOS (10 DÍGITOS)
        # ==========================================
        if not cedula.isdigit() or len(cedula) != 10:
            messages.error(request, "Error de seguridad: La cédula debe tener exactamente 10 dígitos numéricos.")
            return redirect('registro_socio')
            
        if not telefono_personal.isdigit() or len(telefono_personal) != 10:
            messages.error(request, "Error de seguridad: El teléfono debe tener exactamente 10 dígitos numéricos.")
            return redirect('registro_socio')
        # ==========================================

        if User.objects.filter(username=cedula).exists():
            messages.error(request, "Esta cédula ya está registrada.")
            return redirect('registro_socio')

        try:
            fecha_nac = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
            if fecha_nac > date.today():
                messages.error(request, "La fecha de nacimiento no puede ser en el futuro.")
                return redirect('registro_socio')
        except ValueError:
            messages.error(request, "Formato de fecha inválido.")
            return redirect('registro_socio')

        try:
            user = User.objects.create_user(username=cedula, email=email, password=password, first_name=primer_nombre, last_name=primer_apellido)
            tipo_base = TipoSocio.objects.first()
            if not tipo_base:
                user.delete() 
                messages.error(request, "Error crítico: No hay 'Tipos de Socio'.")
                return redirect('registro_socio')
            
            Socio.objects.create(
                idtiposocio=tipo_base, primernombresocio=primer_nombre, segundonombresocio=segundo_nombre,
                primerapellidosocio=primer_apellido, segundoapellidosocio=segundo_apellido, cisocio=cedula,
                fechanacimientosocio=fecha_nac, telefonopersonalsocio=telefono_personal,
                direcciondomiciliosocio=direccion, sexosocio=sexo, estadosocio='Activo'
            )
            login(request, user)
            request.session['preguntar_jugador'] = True
            request.session['user_nombre'] = primer_nombre
            return redirect('inicio')
        except Exception as e:
            if 'user' in locals() and user.id: user.delete() 
            messages.error(request, f"Error en el formulario: {str(e)}")
            return redirect('registro_socio')
    return render(request, 'cuentas/registro_socio.html')

def registro_jugador(request):
    if request.method == 'POST':
        alias = request.POST.get('aliasjugador')
        if request.user.is_authenticated:
            try:
                socio_vinculado = Socio.objects.get(cisocio=request.user.username)
                Jugador.objects.create(idsocio=socio_vinculado, aliasjugador=alias, nombresjugador=socio_vinculado.primernombresocio, cedulaidentidadjugador=socio_vinculado.cisocio, correojugador=request.user.email)
                request.session['user_nombre'] = alias
                messages.success(request, f"¡Perfil de juego activado como '{alias}'!")
                return redirect('inicio')
            except Exception:
                messages.error(request, "Error al vincular el perfil de juego.")
        else:
            nombres, apellidos = request.POST.get('nombresjugador'), request.POST.get('apellidosjugador')
            cedula, correo, password = request.POST.get('cedula'), request.POST.get('correojugador'), request.POST.get('password')

            # ==========================================
            # FIX: ESCUDO DE INTEGRIDAD DE DATOS
            # ==========================================
            if cedula and (not cedula.isdigit() or len(cedula) != 10):
                messages.error(request, "Error de seguridad: La cédula debe tener exactamente 10 dígitos numéricos.")
                return redirect('registro_jugador')
            # ==========================================
            
            if User.objects.filter(username=cedula).exists():
                messages.error(request, "Cédula ya registrada.")
                return redirect('registro_jugador')
            try:
                user = User.objects.create_user(username=cedula, email=correo, password=password, first_name=nombres, last_name=apellidos)
                Jugador.objects.create(aliasjugador=alias, nombresjugador=nombres, apellidosjugador=apellidos, cedulaidentidadjugador=cedula, correojugador=correo)
                login(request, user)
                request.session['user_nombre'] = alias
                messages.success(request, f"¡Bienvenido a la sala de juegos, {alias}!")
                return redirect('inicio')
            except Exception as e:
                if 'user' in locals() and user.id: user.delete()
                messages.error(request, f"Error: {str(e)}")
                return redirect('registro_jugador')
    return render(request, 'cuentas/registro_jugador.html')

# =========================================================
# LÓGICA DE TIENDA Y GENERACIÓN EN TIEMPO REAL
# =========================================================
@login_required
def venta_cartones(request):
    jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
    if not jugador:
        messages.warning(request, "Debes activar tu perfil de juego para entrar a la tienda.")
        return redirect('registro_jugador')

    if jugador.estadocuentajugador != 'Activo':
        messages.error(request, "Tu cuenta de jugador se encuentra suspendida o inactiva. No puedes realizar compras.")
        return redirect('inicio')

    if request.method == 'POST':
        id_bingo = request.POST.get('id_bingo')
        bingo = get_object_or_404(Bingo, idbingo=id_bingo)
        cartones_catalogo_ids = request.POST.getlist('cartones_catalogo')
        cartones_generados_json = request.POST.get('cartones_generados', '[]')
        
        try: cartones_generados = json.loads(cartones_generados_json)
        except Exception: cartones_generados = []

        cantidad_total_compra = len(cartones_catalogo_ids) + len(cartones_generados)

        if cantidad_total_compra == 0:
            messages.error(request, "No seleccionaste ni generaste ningún cartón para comprar.")
            return redirect('venta_cartones')

        cartones_ya_comprados = CartonPartidaBingo.objects.filter(idjugador=jugador, idpartida__idbingo=bingo).values('idcarton').distinct().count()

        

        precio_unitario = bingo.preciocarton
        total_pagar = precio_unitario * cantidad_total_compra

        if jugador.saldocreditojugador < total_pagar:
            messages.error(request, f"Fondos insuficientes. El total es ${total_pagar} y dispones de ${jugador.saldocreditojugador}.")
            return redirect('venta_cartones')

        partidas = PartidaBingo.objects.filter(idbingo=bingo)
        cartones_a_asignar = []

        if cartones_catalogo_ids:
            usados = CartonPartidaBingo.objects.filter(idpartida__in=partidas, idcarton__in=cartones_catalogo_ids).exists()
            if usados:
                messages.error(request, "Oops. Un jugador más rápido compró uno de los cartones de catálogo que elegiste. Vuelve a intentarlo.")
                return redirect('venta_cartones')
            catalogo_validos = Carton.objects.filter(idcarton__in=cartones_catalogo_ids)
            cartones_a_asignar.extend(list(catalogo_validos))

        if cartones_generados:
            nuevos_cartones_db = [Carton(codigocarton=c_data['codigo'], matriznumeros=c_data['matriz'], esmaestro=False) for c_data in cartones_generados]
            Carton.objects.bulk_create(nuevos_cartones_db)
            codigos_creados = [c['codigo'] for c in cartones_generados]
            cartones_temporales = Carton.objects.filter(codigocarton__in=codigos_creados)
            cartones_a_asignar.extend(list(cartones_temporales))

        try:
            jugador.saldocreditojugador -= total_pagar
            jugador.save()

            nuevas_asignaciones = []
            for carton in cartones_a_asignar:
                for partida in partidas:
                    nuevas_asignaciones.append(CartonPartidaBingo(idjugador=jugador, idpartida=partida, idcarton=carton, preciopagado=precio_unitario, estadocarton='Vendido', fechacompra=datetime.now()))
            
            if nuevas_asignaciones:
                CartonPartidaBingo.objects.bulk_create(nuevas_asignaciones)
            
            # ==========================================
            # MAGIA 5: AVISAR A LA TIENDA EN TIEMPO REAL
            # ==========================================
            channel_layer = get_channel_layer()
            for carton in cartones_a_asignar:
                # El grupo de la tienda usa el ID del Bingo maestro
                async_to_sync(channel_layer.group_send)(
                    f'bingo_tienda_{bingo.idbingo}',
                    {
                        'type': 'evento_tienda',
                        'datos': {
                            'evento': 'carton_vendido',
                            'id_carton': carton.idcarton
                        }
                    }
                )
            # ==========================================
            
            messages.success(request, f"¡Adrenalina pura! Tus {cantidad_total_compra} cartones han sido registrados en la base de datos para el evento '{bingo.titulobingo}'.")
            return redirect('venta_cartones')
        
        except Exception as e:
            messages.error(request, f"Fallo crítico en la transacción: {str(e)}")
            return redirect('venta_cartones')

    bingos_disponibles = Bingo.objects.exclude(estadobingo__in=['Finalizado', 'Cancelado']).filter(partidabingo__isnull=False).distinct()
    bingos_data = []
    for b in bingos_disponibles:
        comprados = CartonPartidaBingo.objects.filter(idjugador=jugador, idpartida__idbingo=b).values('idcarton').distinct().count()
        porcentaje_barra = min(int((comprados / 15) * 100), 100)
        usados_ids = CartonPartidaBingo.objects.filter(idpartida__idbingo=b).values_list('idcarton', flat=True)
        catalogo = Carton.objects.filter(esmaestro=True).exclude(idcarton__in=usados_ids)[:12]

        bingos_data.append({'bingo': b, 'comprados': comprados, 'porcentaje': porcentaje_barra, 'catalogo': catalogo})

    contexto = {'jugador': jugador, 'bingos_data': bingos_data}
    return render(request, 'negocio/venta_cartones.html', contexto)


# =========================================================
# VISTA: MI PERFIL Y ESTADO DE CUENTA
# =========================================================
@login_required
def perfil(request):
    user = request.user
    socio = Socio.objects.filter(cisocio=user.username).first()
    jugador = Jugador.objects.filter(cedulaidentidadjugador=user.username).first()

    if request.method == 'POST':
        action = request.POST.get('action')

        try:
            if action == 'actualizar_datos':
                nuevo_correo = request.POST.get('correo')
                if nuevo_correo:
                    user.email = nuevo_correo
                    user.save()
                
                if socio:
                    socio.telefonopersonalsocio = request.POST.get('telefono', socio.telefonopersonalsocio)
                    socio.save()
                    
                if jugador:
                    jugador.aliasjugador = request.POST.get('alias', jugador.aliasjugador)
                    jugador.correojugador = nuevo_correo
                    jugador.save()
                    request.session['user_nombre'] = jugador.aliasjugador
                    
                messages.success(request, "Tus datos de contacto han sido actualizados.")

            elif action == 'actualizar_avatar':
                nueva_foto = request.FILES.get('avatar')
                if nueva_foto:
                    # IMPLEMENTACIÓN NUEVA: Limpieza del servidor
                    actualizar_avatar_perfil(request, socio, jugador, nueva_foto)
                    messages.success(request, "¡Tu foto de perfil luce genial! (Servidor optimizado y limpio).")

            elif action == 'actualizar_password':
                actual = request.POST.get('password_actual')
                nueva = request.POST.get('password_nueva')
                
                if user.check_password(actual):
                    user.set_password(nueva)
                    user.save()
                    update_session_auth_hash(request, user)
                    messages.success(request, "Tu contraseña ha sido cambiada de forma segura.")
                else:
                    messages.error(request, "La contraseña actual no coincide. No se guardaron los cambios.")
            
            elif action == 'ascender_socio':
                # Ahora recibimos TODO del formulario explícitamente
                cedula = request.POST.get('cedula')
                primer_nombre = request.POST.get('primer_nombre')
                segundo_nombre = request.POST.get('segundo_nombre', '')
                primer_apellido = request.POST.get('primer_apellido')
                segundo_apellido = request.POST.get('segundo_apellido')
                telefono = request.POST.get('telefono')
                direccion = request.POST.get('direccion')
                fecha_nacimiento_str = request.POST.get('fecha_nacimiento')
                sexo = request.POST.get('sexo')
                
                try:
                    fecha_nac = datetime.strptime(fecha_nacimiento_str, '%Y-%m-%d').date()
                    tipo_base = TipoSocio.objects.first()
                    
                    # 1. Creamos el Socio con los datos EXACTOS y legales que el usuario llenó
                    nuevo_socio = Socio.objects.create(
                        idtiposocio=tipo_base,
                        primernombresocio=primer_nombre,
                        segundonombresocio=segundo_nombre,
                        primerapellidosocio=primer_apellido,
                        segundoapellidosocio=segundo_apellido,
                        cisocio=cedula,
                        fechanacimientosocio=fecha_nac,
                        telefonopersonalsocio=telefono,
                        direcciondomiciliosocio=direccion,
                        sexosocio=sexo,
                        estadosocio='Activo'
                    )
                    
                    # 2. Actualizamos el User base de Django por si corrigió su cédula o nombres al ascender
                    user.username = cedula
                    user.first_name = primer_nombre
                    user.last_name = primer_apellido
                    user.save()
                    
                    # 3. Vinculamos al jugador
                    jugador.idsocio = nuevo_socio
                    
                    # 4. Limpiamos la redundancia (Vaciamos los datos del jugador como pediste)
                    jugador.nombresjugador = None
                    jugador.apellidosjugador = None
                    
                    jugador.save()
                    
                    messages.success(request, "¡Felicidades! Ahora eres Socio oficial. Tus datos legales han sido registrados con éxito.")
                except Exception as e:
                    messages.error(request, f"Error al procesar la solicitud de socio: {str(e)}")

        except Exception as e:
            messages.error(request, f"Error al actualizar el perfil: {str(e)}")

        return redirect('perfil')

    historial_compras = []
    historial_prestamos = []
    historial_ahorros = []
    
    if jugador:
        historial_compras = CartonPartidaBingo.objects.filter(idjugador=jugador).select_related('idpartida', 'idcarton').order_by('-fechacompra')[:15]
    if socio:
        historial_prestamos = Prestamo.objects.filter(idsocio=socio).order_by('-fechasolicitud')
        historial_ahorros = Ahorro.objects.filter(idsocio=socio).order_by('-fechaahorro')[:15]

    contexto = {
        'socio': socio,
        'jugador': jugador,
        'historial_compras': historial_compras,
        'historial_prestamos': historial_prestamos,
        'historial_ahorros': historial_ahorros,
    }
    return render(request, 'cuentas/perfil.html', contexto)


# =========================================================
# VISTA NUEVA: MIS CARTONES / BÓVEDA DE EVENTOS
# =========================================================
@login_required
def mis_cartones(request):
    try:
        jugador = Jugador.objects.get(correojugador=request.user.email)
    except Jugador.DoesNotExist:
        return redirect('inicio')
        
    cartones_jugador = CartonPartidaBingo.objects.filter(idjugador=jugador).select_related(
        'idcarton', 'idpartida', 'idpartida__idbingo'
    ).order_by('-idpartida__idbingo__fechaprogramadabingo')
    
    # ==============================================================
    # FIX: DEDUPLICADOR DE CARTONES (1 solo cartón visual por Bingo)
    # ==============================================================
    bingos_dict = {}
    for c in cartones_jugador:
        b_id = c.idpartida.idbingo.idbingo
        if b_id not in bingos_dict:
            bingos_dict[b_id] = {
                'bingo': c.idpartida.idbingo,
                'cartones_unicos': {} # Usamos un diccionario para evitar repetidos
            }
        
        # Guardamos el cartón usando su ID como llave. Si ya existe, se ignora.
        carton_id = c.idcarton.idcarton
        if carton_id not in bingos_dict[b_id]['cartones_unicos']:
            bingos_dict[b_id]['cartones_unicos'][carton_id] = c

    # Convertimos de nuevo a lista para la plantilla HTML
    bingos_agrupados = []
    for b_id, data in bingos_dict.items():
        bingos_agrupados.append({
            'bingo': data['bingo'],
            'cartones': list(data['cartones_unicos'].values())
        })
        
    context = {
        'bingos_agrupados': bingos_agrupados,
        'jugador': jugador
    }
    return render(request, 'cuentas/mis_cartones.html', context)

@login_required
def sacar_bola_api(request, id_partida):
    """API para extraer una bola y transmitirla vía WebSockets"""
    if not request.method == 'POST' or not request.user.is_staff:
        return JsonResponse({'error': 'Acceso denegado'}, status=403)

    partida = get_object_or_404(PartidaBingo, idpartidabingo=id_partida)
    
    if partida.estadopartida != 'En Juego':
        return JsonResponse({'error': 'La partida no está en curso'}, status=400)

    # 1. Obtener las bolas ya cantadas
    bolas_str = partida.bolascantadas.replace('B','').replace('I','').replace('N','').replace('G','').replace('O','')
    bolas_llamadas = [int(b.strip()) for b in bolas_str.split(',') if b.strip().isdigit()]

    # 2. Elegir una nueva bola que no haya salido
    bolas_disponibles = [i for i in range(1, 76) if i not in bolas_llamadas]
    if not bolas_disponibles:
        return JsonResponse({'error': 'No hay más bolas disponibles'}, status=400)

    nueva_bola = random.choice(bolas_disponibles)
    bolas_llamadas.append(nueva_bola)

    # 3. Guardar en Base de Datos
    partida.ultimabola = nueva_bola
    partida.bolascantadas = ",".join(map(str, bolas_llamadas))
    partida.save()

    # 4. Magia en Tiempo Real: Enviar evento al grupo WebSocket de la partida
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        f'bingo_partida_{id_partida}',
        {
            'type': 'evento_partida',
            'datos': {
                'evento': 'nueva_bola',
                'numero': nueva_bola
            }
        }
    )

    return JsonResponse({'status': 'ok', 'bola_extraida': nueva_bola})

@login_required
def descargar_cartones_pdf(request, id_bingo):
    if request.method == 'POST':
        jugador = Jugador.objects.filter(cedulaidentidadjugador=request.user.username).first()
        if not jugador:
            messages.error(request, "Perfil no encontrado.")
            return redirect('mis_cartones')

        cartones_ids = request.POST.getlist('cartones_seleccionados')
        if not cartones_ids:
            messages.warning(request, "No seleccionaste ningún cartón para imprimir.")
            return redirect('mis_cartones')

        bingo = get_object_or_404(Bingo, idbingo=id_bingo)
        cartones_asignados = CartonPartidaBingo.objects.filter(
            idjugador=jugador, 
            idpartida__idbingo=bingo,
            idcarton__in=cartones_ids
        ).select_related('idcarton')

        # 1. DEDUPLICACIÓN: Evitamos que el mismo cartón salga varias veces si hay varias rondas
        cartones_unicos = {}
        for asig in cartones_asignados:
            if asig.idcarton.idcarton not in cartones_unicos:
                matriz = asig.idcarton.matriznumeros
                if isinstance(matriz, str):
                    try: matriz = json.loads(matriz.replace("'", '"'))
                    except: continue
                
                if isinstance(matriz, dict):
                    try:
                        filas = []
                        for i in range(5):
                            filas.append([matriz['B'][i], matriz['I'][i], matriz['N'][i], matriz['G'][i], matriz['O'][i]])
                        
                        cartones_unicos[asig.idcarton.idcarton] = {
                            'codigo': asig.idcarton.codigocarton,
                            'filas': filas
                        }
                    except Exception as e: print(e)

        cartones_procesados = list(cartones_unicos.values())

        # 2. Renderizar y generar PDF
        template = get_template('cuentas/cartones_pdf.html')
        context = {'bingo': bingo, 'jugador': jugador, 'cartones': cartones_procesados}
        html = template.render(context)

        response = HttpResponse(content_type='application/pdf')
        # MAGIA: Cambiamos 'attachment' por 'inline' para que el navegador lo PREVISUALICE
        response['Content-Disposition'] = f'inline; filename="Mis_Cartones_{bingo.idbingo}_{jugador.aliasjugador}.pdf"'
        
        pisa_status = pisa.CreatePDF(html, dest=response)
        
        if pisa_status.err:
            return HttpResponse('Tuvimos errores generando tu documento PDF', status=500)
        return response
    
    return redirect('mis_cartones')

@login_required
def reporte_socios_puntuales(request):
    if not request.user.is_staff: return redirect('inicio')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Socios Estrella"
    
    ws.append(['Cédula', 'Socio', 'Teléfono', 'Tipo de Socio', 'Historial de Aportes', 'Calificación'])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="312E81", fill_type="solid") # Azul corporativo
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    socios = Socio.objects.filter(estadosocio='Activo').select_related('idtiposocio')
    for s in socios:
        aportes = AporteSemanal.objects.filter(idsocio=s)
        total_aportes = aportes.count()
        aportes_al_dia = aportes.filter(estadoaporte='Al Dia').count()
        
        clasificacion = "Sin Historial"
        if total_aportes > 0:
            porcentaje = (aportes_al_dia / total_aportes) * 100
            if porcentaje == 100: clasificacion = "🌟 EXCELENTE (Aplica Descuento)"
            elif porcentaje >= 80: clasificacion = "👍 BUENO (Cumplido)"
            elif porcentaje >= 50: clasificacion = "⚠️ REGULAR (Alerta)"
            else: clasificacion = "❌ MOROSO (Riesgo Alto)"

        ws.append([
            s.cisocio,
            f"{s.primernombresocio} {s.primerapellidosocio}",
            s.telefonopersonalsocio,
            s.idtiposocio.nombretiposocio if s.idtiposocio else "No Definido",
            f"{aportes_al_dia} de {total_aportes} Al Día",
            clasificacion
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Socios_Estrella_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def reporte_liquidacion_bingo(request, id_bingo):
    if not request.user.is_staff: return redirect('inicio')
    
    bingo = get_object_or_404(Bingo, idbingo=id_bingo)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Liquidación de Bingo"

    ws.append(['Concepto', 'Detalle', 'Monto Total'])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E1B4B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Obtención de métricas mediante agregaciones en la base de datos
    cartones_vendidos = CartonPartidaBingo.objects.filter(idpartida__idbingo=bingo).values('idcarton').distinct().count()
    ingresos_totales = cartones_vendidos * bingo.preciocarton
    
    # Manejo adaptativo de campos por si la base varía el nombre del premio
    premios_entregados = 0
    try: premios_entregados = PartidaBingo.objects.filter(idbingo=bingo).aggregate(total=Sum('valorpremio'))['total'] or 0
    except:
        try: premios_entregados = PartidaBingo.objects.filter(idbingo=bingo).aggregate(total=Sum('valorefectivo'))['total'] or 0
        except: pass
        
    utilidad_neta = ingresos_totales - premios_entregados

    ws.append(['INGRESOS', f'Recaudación por Cartones ({cartones_vendidos} x ${bingo.preciocarton})', ingresos_totales])
    ws.append(['EGRESOS', 'Total Premios en Efectivo Entregados en Rondas', -premios_entregados])
    ws.append(['UTILIDAD LÍQUIDA', 'Balance Neto de la Cooperativa', utilidad_neta])
    
    ws[4][2].font = Font(bold=True, color="008000" if utilidad_neta >= 0 else "FF0000")
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Liquidacion_{bingo.idbingo}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def reporte_cartera_prestamos(request):
    if not request.user.is_staff: return redirect('inicio')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cartera de Créditos"
    
    ws.append(['Cédula', 'Socio', 'Monto Solicitado', 'Total a Pagar', 'Saldo Pendiente', 'Estado'])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="B91C1C", fill_type="solid") # Rojo analítico financiero
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    prestamos = Prestamo.objects.all().select_related('idsocio').order_by('-fechasolicitud')
    for p in prestamos:
        ws.append([
            p.idsocio.cisocio if p.idsocio else "N/A",
            f"{p.idsocio.primernombresocio} {p.idsocio.primerapellidosocio}" if p.idsocio else "Externo",
            float(p.montoprestamosolicitado or 0),
            float(p.montototalpagar or 0),
            float(p.saldopendiente or 0),
            p.estadoprestamo
        ])
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col[0].column)].width = max(max_len + 3, 12)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Cartera_Creditos_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def reporte_caja_semanal_pdf(request):
    if not request.user.is_staff: return redirect('inicio')
    
    aportes = AporteSemanal.objects.all().select_related('idsocio').order_by('-fechaplanificadadada', 'idsocio__primerapellidosocio')
    total_recaudado = aportes.filter(estadoaporte='Al Dia').aggregate(total=Sum('montoaportesemanal'))['total'] or 0
    total_pendiente = aportes.filter(estadoaporte='Pendiente').aggregate(total=Sum('montoaportesemanal'))['total'] or 0

    template = get_template('administrador/reporte_caja_pdf.html')
    context = {
        'aportes': aportes,
        'total_recaudado': total_recaudado,
        'total_pendiente': total_pendiente,
        'fecha_reporte': timezone.now()
    }
    html = template.render(context)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="Cierre_Caja_Semanal.pdf"'
    pisa.CreatePDF(html, dest=response)
    return response

def control_aportes(request):
    # 1. Filtro por el Bingo seleccionado o el más reciente por defecto
    id_bingo = request.GET.get('bingo_id')
    if id_bingo:
        bingo_seleccionado = Bingo.objects.filter(idbingo=id_bingo).first()
    else:
        bingo_seleccionado = Bingo.objects.filter(estadobingo='En Curso').first() or Bingo.objects.order_by('-fechaprogramadabingo').first()

    if not bingo_seleccionado:
        return render(request, 'negocio/control_aportes.html', {'error': 'No hay eventos de bingo creados.'})

    # 2. Traer todos los socios activos para las filas
    socios = Socio.objects.filter(estadosocio='Activo').order_by('primerapellidosocio', 'primernombresocio')

    # 3. Obtener los aportes existentes de este bingo
    aportes = AporteSemanal.objects.filter(idbingo=bingo_seleccionado).select_related('idsocio')

    # 4. Determinar cuántas semanas mostraremos (basado en lo que ya hay guardado o un rango por defecto)
    semanas_query = aportes.values_list('numerosemana', flat=True).distinct().order_by('numerosemana')
    semanas = list(semanas_query) if semanas_query.exists() else list(range(1, 6)) # Muestra mínimo 5 semanas si está en blanco

    # 5. Armar la estructura matricial { socio_id: { info_socio, semanas: { num_sem: aporte_obj }, total } }
    matriz_socios = {}
    for socio in socios:
        matriz_socios[socio.idsocio] = {
            'objeto_socio': socio,
            'semanas_data': {sem: None for sem in semanas},
            'total_acumulado': Decimal('0.00'),
            'tiene_atrasos': False
        }

    # 6. Llenar las celdas con los registros reales de la base de datos
    for aporte in aportes:
        s_id = aporte.idsocio_id
        if s_id in matriz_socios:
            # Si la semana del aporte entra en el rango visible
            if aporte.numerosemana in matriz_socios[s_id]['semanas_data']:
                matriz_socios[s_id]['semanas_data'][aporte.numerosemana] = {
                    'monto': aporte.montoaporte,
                    'estado': aporte.estadoaporte,
                    'id_aporte': aporte.idaporte
                }
                # Sumar al total si el aporte fue validado/al día
                if aporte.estadoaporte == 'Al Dia':
                    matriz_socios[s_id]['total_acumulado'] += aporte.montoaporte
                elif aporte.estadoaporte == 'Atrasado':
                    matriz_socios[s_id]['tiene_atrasos'] = True

    context = {
        'bingo_seleccionado': bingo_seleccionado,
        'todos_los_bingos': Bingo.objects.all().order_by('-fechaprogramadabingo'),
        'semanas': semanas,
        'matriz_socios': matriz_socios.values(), # Pasamos los renglones limpios
    }
    return render(request, 'dashboard.html', context)